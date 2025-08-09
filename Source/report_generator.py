import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import logging

def _apply_excel_styles(ws):
    """Excel 시트에 공통 스타일을 적용하는 헬퍼 함수"""
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 헤더 행에 스타일 적용
    for cell in ws[1]:
        cell.fill = header_fill
        
    # 헤더 열(인덱스)에 스타일 적용
    for row in ws.iter_rows(min_row=2):
        row[0].fill = header_fill

    # 열 너비 자동 조정
    for column_cells in ws.columns:
        # 셀 값의 길이를 계산할 때 NoneType을 고려하여 str로 변환
        length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 4


def generate_trip_report(config):
    """
    (최종 수정) 차종별 하위 폴더를 모두 탐색하고, 새로운 파일명 규칙에 따라
    단말기별/월별 Trip 개수 리포트와 요약 리포트를 생성합니다.
    """
    trip_folder_path = config.PATHS["output_trip"]
    report_output_path = config.PATHS["output_report"]
    
    all_trip_files = []
    for root, _, files in os.walk(trip_folder_path):
        for file in files:
            if file.endswith('.csv'):
                car_model = os.path.basename(root)
                all_trip_files.append((car_model, file))

    if not all_trip_files:
        logging.warning(f"리포트할 Trip 파일이 '{trip_folder_path}'에 없습니다.")
        return

    records = []
    for car_model, file in all_trip_files:
        try:
            parts = file.replace('.csv', '').split('_')
            
            device_no = None
            year_month = None
            has_altitude = False

            if file.startswith('Trip_altitude_'):
                if len(parts) >= 5:
                    device_no = parts[2]
                    year_month = parts[3]
                    has_altitude = True
                else:
                    logging.warning(f"파일명 형식이 예상과 다릅니다 (altitude): {file}")
                    continue
            elif file.startswith('Trip_'):
                if len(parts) >= 4:
                    device_no = parts[1]
                    year_month = parts[2]
                else:
                    logging.warning(f"파일명 형식이 예상과 다릅니다: {file}")
                    continue
            else:
                logging.warning(f"지원하지 않는 파일명 형식입니다: {file}")
                continue
            
            records.append({
                "차종": car_model,
                "단말기번호": device_no,
                "연월": year_month,
                "altitude_유무": has_altitude # ✅ altitude 정보 기록
            })
        except IndexError:
            logging.warning(f"파일명 분석 중 오류 발생: {file}")
            continue
            
    if not records:
        logging.warning("분석할 유효한 Trip 레코드가 없습니다.")
        return
        
    df = pd.DataFrame(records)
    
    # --- 1. 기존 리포트 (단말기별/월별) 생성 로직 ---
    report_df = df.groupby(["차종", "단말기번호", "연월"]).size().reset_index(name='Trip 수')
    pivot_df = report_df.pivot_table(
        index=['차종', '단말기번호'], columns='연월', values='Trip 수',
        fill_value=0, aggfunc='sum'
    ).sort_index()
    pivot_df['총 Trip'] = pivot_df.sum(axis=1)

    # 2-1. 차종별 Trip 분석 (총 Trip 수, Altitude 포함 Trip 수)
    summary_by_car = df.groupby('차종').agg(
        총_Trip_수=('차종', 'size'),
        Altitude_Trip_수=('altitude_유무', 'sum')
    ).astype(int) # 정수형으로 변환
    summary_by_car['Altitude_비율(%)'] = round((summary_by_car['Altitude_Trip_수'] / summary_by_car['총_Trip_수']) * 100, 1)

    # 2-2. 차종별/월별 Trip 개수 요약
    summary_by_month = df.groupby(['차종', '연월']).size().unstack(fill_value=0)


    # --- 3. Excel 파일로 저장 ---
    output_excel_file = report_output_path / 'Trip_report.xlsx'
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        # 첫 번째 시트: 기존 리포트 저장
        pivot_df.to_excel(writer, sheet_name='단말기별_Trip_현황', merge_cells=False)
        _apply_excel_styles(writer.sheets['단말기별_Trip_현황'])
        
        # 두 번째 시트: 신규 요약 리포트 저장
        # 제목을 추가하기 위해 startrow=1 부터 데이터프레임 작성
        summary_by_car.to_excel(writer, sheet_name='Trip_요약', startrow=1)
        ws_summary = writer.sheets['Trip_요약']
        ws_summary['A1'] = "차종별 Trip 분석"

        # 두 번째 테이블을 아래에 이어서 작성 (3칸 띄우기)
        start_row_for_month_summary = len(summary_by_car) + 5
        summary_by_month.to_excel(writer, sheet_name='Trip_요약', startrow=start_row_for_month_summary)
        ws_summary.cell(row=start_row_for_month_summary, column=1, value="차종별/월별 Trip 분석")

        # 새로 추가된 'Trip_요약' 시트에도 스타일 적용
        _apply_excel_styles(ws_summary)


    logging.info(f"🎉 종합 리포트가 '{output_excel_file}'에 성공적으로 저장되었습니다.")